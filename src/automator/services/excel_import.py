"""Tolerant Excel importer for suppliers and buying companies.

Only the Excel reading touches IO (openpyxl); everything else is pure and
testable with plain dict rows. Headers are matched flexibly (accents, case,
dots ignored), extra alias columns are folded in, and each row is validated
independently so one bad row never aborts the whole import.
"""

from __future__ import annotations

import re
from collections.abc import Callable, Iterable
from dataclasses import dataclass
from pathlib import Path
from typing import Generic, TypeVar

from openpyxl import load_workbook
from pydantic import ValidationError

from automator.config import SocietyMapping
from automator.domain.names import normalize_name
from automator.domain.suppliers import Supplier

T = TypeVar("T")

_NON_ALNUM = re.compile(r"[^a-z0-9]")
_HEADER_ALIASES: dict[str, tuple[str, ...]] = {
    "cuit": ("cuit", "cuitcuil", "cuil"),
    "razon_social": ("razonsocial", "razon"),
    "nombre_fantasia": ("nombredefantasia", "nombrefantasia", "fantasia"),
}
_ALIAS_MARKERS = ("alias", "variant", "variacion")
_FIRST_DATA_ROW = 2  # Row 1 is the header; spreadsheet rows are 1-based.


class MissingColumnError(Exception):
    """A required column (CUIT or Razon Social) is absent from the sheet."""


class ExcelReadError(Exception):
    """The Excel file could not be opened or read (corrupt, wrong format, locked)."""


@dataclass(frozen=True)
class ImportReport(Generic[T]):
    """Outcome of an import: the entities created and the rows that were rejected."""

    created: list[T]
    invalid: list[tuple[int, str]]  # (spreadsheet row number, reason)


@dataclass(frozen=True)
class _Fields:
    cuit: str
    razon_social: str
    nombre_fantasia: str | None
    aliases: tuple[str, ...]


def read_rows(path: Path) -> list[dict[str, str]]:
    """Read the first sheet into a list of header-keyed rows (the only IO)."""
    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:
        raise ExcelReadError(str(exc)) from exc
    try:
        rows = workbook.active.iter_rows(values_only=True)
        headers = [_cell_str(value) for value in next(rows, ())]
        return [record for row in rows if (record := _row_record(headers, row))]
    finally:
        workbook.close()


def map_columns(headers: Iterable[str]) -> dict[str, str]:
    """Map each logical field to the actual header that names it, if present."""
    resolved: dict[str, str] = {}
    for header in headers:
        key = _header_key(header)
        for field_name, variants in _HEADER_ALIASES.items():
            if key in variants and field_name not in resolved:
                resolved[field_name] = header
    return resolved


def parse_suppliers(rows: list[dict[str, str]]) -> ImportReport[Supplier]:
    return _parse(rows, _build_supplier, _merge_supplier)


def parse_societies(rows: list[dict[str, str]]) -> ImportReport[SocietyMapping]:
    return _parse(rows, _build_society, _merge_society)


def _parse(
    rows: list[dict[str, str]],
    build: Callable[[_Fields], T],
    merge: Callable[[dict[str, T], T], None],
) -> ImportReport[T]:
    columns, alias_headers = _columns(rows)
    created: dict[str, T] = {}
    invalid: list[tuple[int, str]] = []
    for offset, row in enumerate(rows):
        try:
            entity = build(_extract(row, columns, alias_headers))
        except (ValidationError, ValueError) as exc:
            invalid.append((offset + _FIRST_DATA_ROW, _reason(exc)))
            continue
        merge(created, entity)
    return ImportReport(created=list(created.values()), invalid=invalid)


def _columns(rows: list[dict[str, str]]) -> tuple[dict[str, str], list[str]]:
    headers = list(dict.fromkeys(key for row in rows for key in row))
    columns = map_columns(headers)
    if "cuit" not in columns or "razon_social" not in columns:
        raise MissingColumnError("El Excel debe tener columnas de CUIT y Razon Social.")
    return columns, _alias_headers(headers)


def _extract(row: dict[str, str], columns: dict[str, str], alias_headers: list[str]) -> _Fields:
    fantasia = row.get(columns["nombre_fantasia"], "").strip() if "nombre_fantasia" in columns else ""
    aliases = tuple(alias for header in alias_headers if (alias := row.get(header, "").strip()))
    return _Fields(
        cuit=row.get(columns["cuit"], "").strip(),
        razon_social=row.get(columns["razon_social"], "").strip(),
        nombre_fantasia=fantasia or None,
        aliases=aliases,
    )


def _build_supplier(fields: _Fields) -> Supplier:
    return Supplier(
        cuit=fields.cuit,
        razon_social=fields.razon_social,
        nombre_fantasia=fields.nombre_fantasia,
        extra_aliases=fields.aliases,
    )


def _build_society(fields: _Fields) -> SocietyMapping:
    return SocietyMapping(
        cuit=fields.cuit,
        name=fields.razon_social,
        nombre_fantasia=fields.nombre_fantasia,
        aliases=fields.aliases,
    )


def _merge_supplier(created: dict[str, Supplier], supplier: Supplier) -> None:
    existing = created.get(supplier.cuit)
    if existing is None:
        created[supplier.cuit] = supplier
        return
    created[supplier.cuit] = existing.model_copy(
        update={
            "extra_aliases": _dedupe(existing.extra_aliases, supplier.extra_aliases),
            "nombre_fantasia": existing.nombre_fantasia or supplier.nombre_fantasia,
        }
    )


def _merge_society(created: dict[str, SocietyMapping], society: SocietyMapping) -> None:
    existing = created.get(society.cuit)
    if existing is None:
        created[society.cuit] = society
        return
    created[society.cuit] = existing.model_copy(
        update={
            "aliases": _dedupe(existing.aliases, society.aliases),
            "nombre_fantasia": existing.nombre_fantasia or society.nombre_fantasia,
        }
    )


def _dedupe(*groups: tuple[str, ...]) -> tuple[str, ...]:
    return tuple(dict.fromkeys(alias for group in groups for alias in group))


def _reason(exc: ValidationError | ValueError) -> str:
    if isinstance(exc, ValidationError):
        return str(exc.errors()[0]["msg"]).removeprefix("Value error, ")
    return str(exc)


def _header_key(text: str) -> str:
    return _NON_ALNUM.sub("", normalize_name(text))


def _alias_headers(headers: Iterable[str]) -> list[str]:
    return [header for header in headers if any(marker in _header_key(header) for marker in _ALIAS_MARKERS)]


def _row_record(headers: list[str], row: tuple[object, ...]) -> dict[str, str]:
    record = {header: _cell_str(value) for header, value in zip(headers, row, strict=False) if header}
    return record if any(record.values()) else {}


def _cell_str(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()
